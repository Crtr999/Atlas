#!/usr/bin/env python3
"""
CBC Atlas Executive Report Generator

Produces a formatted, colour-coded Excel workbook with four sheets:

    1. Executive Summary  — high-level KPIs and top-5 breakdowns
    2. Full Dataset        — every case, colour-coded by status, filterable
    3. Denial Hotspots     — counties and judges ranked by denial rate
    4. State Breakdown     — per-state outcome table

Usage:
    python generate_executive_report.py                  # saves to ./reports/
    python generate_executive_report.py --output /tmp    # custom output dir

Returns the filepath of the generated report (for Slack upload or sharing).
"""

import argparse
import os
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

load_dotenv()
EXCEL_PATH = os.environ.get("EXCEL_PATH", "./CBC_Settlement_Funding_Master_v4.xlsx")

# ── Palette ─────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy
C_HEADER_FONT = "FFFFFF"   # white
C_SUBHEAD_BG  = "2E75B6"   # medium blue
C_SUBHEAD_FONT= "FFFFFF"
C_TITLE_FONT  = "1F4E79"
C_BAND        = "EBF3FB"   # very light blue row bands

C_APPROVED_BG = "E2EFDA"   # light green
C_DENIED_BG   = "FFE0E0"   # light red
C_DISMISSED_BG= "FFF2CC"   # light amber
C_OTHER_BG    = "F2F2F2"   # light gray

C_KPI_BG      = "D6E4F0"   # light blue for KPI cells
C_ALERT_BG    = "FFD7D7"   # pink for high-denial alert rows
C_GRAY_BORDER = "BFBFBF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)


def _thin_border() -> Border:
    s = Side(style="thin", color=C_GRAY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _normalize_status(series: pd.Series) -> pd.Series:
    s = series.fillna("").astype(str).str.lower()
    result = pd.Series("Other", index=series.index)
    result[s.str.contains("approved|granted")] = "Approved"
    result[s.str.contains("denied")] = "Denied"
    result[s.str.contains("dismiss")] = "Dismissed"
    return result


def _status_fill(status: str) -> PatternFill:
    return _fill({
        "Approved":  C_APPROVED_BG,
        "Denied":    C_DENIED_BG,
        "Dismissed": C_DISMISSED_BG,
    }.get(status, C_OTHER_BG))


# ── Sheet 1 — Executive Summary ─────────────────────────────────────

def _build_summary(wb: Workbook, df: pd.DataFrame, gen_date: str):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "CBC Settlement Funding  —  Executive Summary"
    c.font = _font(bold=True, size=18, color=C_TITLE_FONT)
    c.alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Generated {gen_date}     |     {len(df):,} total cases in dataset"
    c.font = _font(size=10, italic=True, color="595959")
    c.alignment = _center()
    ws.row_dimensions[2].height = 16

    df["_status"] = _normalize_status(df["Status"])

    total     = len(df)
    approved  = int((df["_status"] == "Approved").sum())
    denied    = int((df["_status"] == "Denied").sum())
    dismissed = int((df["_status"] == "Dismissed").sum())
    other     = total - approved - denied - dismissed

    # ── KPI block ────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    _header_row(ws, 4, ["Metric", "Count", "Rate"], col_start=1, bg=C_HEADER_BG)

    kpi_rows = [
        ("Total Cases",              total,     "—"),
        ("Funded / Approved",        approved,  f"{approved/total:.1%}" if total else "—"),
        ("Denied",                   denied,    f"{denied/total:.1%}"   if total else "—"),
        ("Dismissed",                dismissed, f"{dismissed/total:.1%}" if total else "—"),
    ]
    if other > 0:
        kpi_rows.append(("Other / Unknown", other, f"{other/total:.1%}" if total else "—"))

    for r_offset, (label, count, rate) in enumerate(kpi_rows):
        row = 5 + r_offset
        fills = {
            "Funded / Approved": C_APPROVED_BG,
            "Denied":            C_DENIED_BG,
            "Dismissed":         C_DISMISSED_BG,
        }
        row_fill = _fill(fills.get(label, C_KPI_BG))
        for col, val in [(1, label), (2, count), (3, rate)]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = row_fill
            c.font = _font(bold=(col == 1), size=11)
            c.border = _thin_border()
            c.alignment = _center() if col > 1 else _left()

    next_row = 5 + len(kpi_rows) + 2

    # ── Top 5 states by volume ────────────────────────────────────────
    state_counts = (
        df.groupby("State")
        .size()
        .sort_values(ascending=False)
        .head(5)
        .reset_index(name="Cases")
    )
    _section_header(ws, next_row, "Top 5 States by Case Volume", col_span=3)
    next_row += 1
    _header_row(ws, next_row, ["State", "Cases", "Approval Rate"], bg=C_SUBHEAD_BG)
    next_row += 1
    for _, srow in state_counts.iterrows():
        state_df = df[df["State"] == srow["State"]]
        state_approved = int((state_df["_status"] == "Approved").sum())
        n = len(state_df)
        rate = f"{state_approved/n:.1%}" if n else "—"
        for col, val in [(1, srow["State"]), (2, srow["Cases"]), (3, rate)]:
            c = ws.cell(row=next_row, column=col, value=val)
            c.border = _thin_border()
            c.alignment = _center()
            c.font = _font(size=10)
            if next_row % 2 == 0:
                c.fill = _fill(C_BAND)
        next_row += 1

    next_row += 1

    # ── Top denial jurisdictions ──────────────────────────────────────
    denial_rows = []
    for (state, county), grp in df.groupby(["State", "County"]):
        if len(grp) < 3:
            continue
        denied = int((grp["_status"] == "Denied").sum())
        if denied == 0:
            continue
        denial_rows.append({
            "Jurisdiction": f"{county}, {state}",
            "Cases": len(grp),
            "Denied": denied,
            "Denial Rate": denied / len(grp),
        })
    denial_rows.sort(key=lambda x: x["Denial Rate"], reverse=True)

    _section_header(ws, next_row, "Top 5 Denial Hotspots (min 3 cases)", col_span=4)
    next_row += 1
    _header_row(ws, next_row, ["Jurisdiction", "Cases", "Denied", "Denial Rate"], bg=C_SUBHEAD_BG)
    next_row += 1
    for row_data in denial_rows[:5]:
        rate = row_data["Denial Rate"]
        row_fill = _fill(C_ALERT_BG) if rate >= 0.66 else _fill(C_BAND if next_row % 2 == 0 else "FFFFFF")
        for col, key in enumerate(["Jurisdiction", "Cases", "Denied"], start=1):
            c = ws.cell(row=next_row, column=col, value=row_data[key])
            c.fill = row_fill
            c.border = _thin_border()
            c.alignment = _center() if col > 1 else _left()
            c.font = _font(size=10)
        c = ws.cell(row=next_row, column=4, value=f"{rate:.1%}")
        c.fill = row_fill
        c.border = _thin_border()
        c.alignment = _center()
        c.font = _font(bold=True, size=10)
        next_row += 1

    if not denial_rows:
        ws.cell(row=next_row, column=1, value="No jurisdictions with 3+ cases and denials found.")
        next_row += 1

    _set_col_widths(ws, {"A": 36, "B": 12, "C": 12, "D": 14, "E": 14, "F": 14, "G": 14})


# ── Sheet 2 — Full Dataset ───────────────────────────────────────────

def _build_full_dataset(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Full Dataset")

    COLS = ["Client_Name", "State", "County", "Judge",
            "Case_Number", "Court_Date", "Status", "Notes"]
    HEADERS = ["Client Name", "State", "County", "Judge",
               "Case #", "Date", "Status", "Notes"]

    # Header row
    ws.row_dimensions[1].height = 18
    for col_idx, label in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = _fill(C_HEADER_BG)
        c.font = _font(bold=True, color=C_HEADER_FONT, size=11)
        c.alignment = _center()
        c.border = _thin_border()

    # Data rows
    df["_status"] = _normalize_status(df["Status"])
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        status = row.get("_status", "Other")
        row_fill = _status_fill(status)
        for col_idx, col in enumerate(COLS, start=1):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            c = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            c.fill = row_fill
            c.font = _font(size=10)
            c.alignment = _left() if col_idx in (1, 4, 8) else _center()
            c.border = _thin_border()

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # Freeze header
    ws.freeze_panes = "A2"

    _set_col_widths(ws, {
        "A": 36, "B": 7, "C": 20, "D": 22,
        "E": 14, "F": 12, "G": 12, "H": 40,
    })


# ── Sheet 3 — Denial Hotspots ────────────────────────────────────────

def _build_denial_hotspots(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Denial Hotspots")
    ws.sheet_view.showGridLines = False

    df["_status"] = _normalize_status(df["Status"])
    row = 1

    # ── By county ────────────────────────────────────────────────────
    _section_header(ws, row, "Counties Ranked by Denial Rate  (min 3 cases)", col_span=5)
    row += 1
    _header_row(ws, row, ["County", "State", "Total Cases", "Denied", "Denial Rate"], bg=C_HEADER_BG)
    row += 1

    county_rows = []
    for (state, county), grp in df.groupby(["State", "County"]):
        if len(grp) < 3:
            continue
        denied = int((grp["_status"] == "Denied").sum())
        if denied == 0:
            continue
        county_rows.append((county, state, len(grp), denied, denied / len(grp)))
    county_rows.sort(key=lambda x: x[4], reverse=True)

    for i, (county, state, total, denied, rate) in enumerate(county_rows):
        bg = C_ALERT_BG if rate >= 0.5 else (C_BAND if i % 2 == 0 else "FFFFFF")
        vals = [county, state, total, denied, f"{rate:.1%}"]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(col_idx == 5), size=10)
            c.border = _thin_border()
            c.alignment = _left() if col_idx == 1 else _center()
        row += 1

    if not county_rows:
        ws.cell(row=row, column=1,
                value="No counties with 3+ cases and at least one denial.").font = _font(italic=True)
        row += 1

    row += 2

    # ── By judge ─────────────────────────────────────────────────────
    _section_header(ws, row, "Judges Ranked by Denial Rate  (min 3 cases)", col_span=5)
    row += 1
    _header_row(ws, row, ["Judge", "States", "Total Cases", "Denied", "Denial Rate"], bg=C_HEADER_BG)
    row += 1

    judge_rows = []
    for judge, grp in df.groupby("Judge"):
        judge_str = str(judge).strip()
        if not judge_str or judge_str.lower() == "nan" or len(grp) < 3:
            continue
        denied = int((grp["_status"] == "Denied").sum())
        if denied == 0:
            continue
        states = ", ".join(sorted(grp["State"].dropna().astype(str).unique()))
        judge_rows.append((judge_str, states, len(grp), denied, denied / len(grp)))
    judge_rows.sort(key=lambda x: x[4], reverse=True)

    for i, (judge, states, total, denied, rate) in enumerate(judge_rows):
        bg = C_ALERT_BG if rate >= 0.5 else (C_BAND if i % 2 == 0 else "FFFFFF")
        vals = [judge, states, total, denied, f"{rate:.1%}"]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = _fill(bg)
            c.font = _font(bold=(col_idx == 5), size=10)
            c.border = _thin_border()
            c.alignment = _left() if col_idx <= 2 else _center()
        row += 1

    if not judge_rows:
        ws.cell(row=row, column=1,
                value="No judges with 3+ cases and at least one denial.").font = _font(italic=True)

    ws.freeze_panes = "A3"
    _set_col_widths(ws, {"A": 30, "B": 20, "C": 13, "D": 10, "E": 13})


# ── Sheet 4 — State Breakdown ────────────────────────────────────────

def _build_state_breakdown(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("State Breakdown")

    df["_status"] = _normalize_status(df["Status"])
    HEADERS = ["State", "Total Cases", "Approved", "Approval %",
               "Denied", "Denial %", "Dismissed", "Other"]
    _header_row(ws, 1, HEADERS, bg=C_HEADER_BG)
    ws.freeze_panes = "A2"

    rows = []
    for state, grp in df.groupby("State"):
        total     = len(grp)
        approved  = int((grp["_status"] == "Approved").sum())
        denied    = int((grp["_status"] == "Denied").sum())
        dismissed = int((grp["_status"] == "Dismissed").sum())
        other     = total - approved - denied - dismissed
        rows.append((str(state), total, approved, approved / total,
                     denied, denied / total, dismissed, other))
    rows.sort(key=lambda x: x[1], reverse=True)

    for r_idx, row_data in enumerate(rows, start=2):
        state, total, appr, appr_rate, den, den_rate, dis, oth = row_data
        bg = C_BAND if r_idx % 2 == 0 else "FFFFFF"
        if den_rate >= 0.5:
            bg = C_ALERT_BG
        elif appr_rate >= 0.8:
            bg = C_APPROVED_BG

        vals = [state, total, appr, f"{appr_rate:.1%}",
                den, f"{den_rate:.1%}", dis, oth]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.border = _thin_border()
            c.alignment = _center()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    _set_col_widths(ws, {
        "A": 8, "B": 13, "C": 12, "D": 12,
        "E": 10, "F": 10, "G": 12, "H": 10,
    })


# ── Shared helpers ───────────────────────────────────────────────────

def _header_row(ws, row: int, labels: list, col_start: int = 1, bg: str = C_HEADER_BG):
    ws.row_dimensions[row].height = 18
    for col_idx, label in enumerate(labels, start=col_start):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=C_HEADER_FONT, size=11)
        c.alignment = _center()
        c.border = _thin_border()


def _section_header(ws, row: int, text: str, col_span: int = 5):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=col_span,
    )
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(C_SUBHEAD_BG)
    c.font = _font(bold=True, color=C_SUBHEAD_FONT, size=12)
    c.alignment = _left()
    ws.row_dimensions[row].height = 20


# ── Public entry point ───────────────────────────────────────────────

def generate_report(
    excel_path: str = EXCEL_PATH,
    output_dir: str = "./reports",
) -> str:
    """
    Generate the executive report and return the filepath of the
    saved .xlsx file.
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    df = pd.read_excel(path, sheet_name="Cases")
    if df.empty:
        raise ValueError("Cases sheet is empty — nothing to report.")

    gen_date = date.today().strftime("%B %d, %Y")
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    filename = f"CBC_Executive_Report_{date.today().isoformat()}.xlsx"
    out_path = out_dir / filename

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    _build_summary(wb, df.copy(), gen_date)
    _build_full_dataset(wb, df.copy())
    _build_denial_hotspots(wb, df.copy())
    _build_state_breakdown(wb, df.copy())

    wb.save(str(out_path))
    return str(out_path)


# ── CLI ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate CBC Atlas executive Excel report"
    )
    parser.add_argument("--excel", default=EXCEL_PATH,
                        help="Path to Excel workbook")
    parser.add_argument("--output", default="./reports",
                        help="Output directory (default: ./reports)")
    args = parser.parse_args()

    print("Generating executive report...")
    out_path = generate_report(excel_path=args.excel, output_dir=args.output)
    print(f"Report saved: {out_path}")


if __name__ == "__main__":
    main()
